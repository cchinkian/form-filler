import json
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from openpyxl import Workbook
from reportlab.pdfgen import canvas
import pypdf

import excel_reader
import catalog
import package_engine


def _make_pdf(path: Path, title: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    c = canvas.Canvas(str(path), pagesize=(595, 842))
    c.drawString(50, 800, title)
    c.drawString(50, 760, "Name:")
    c.drawString(50, 730, "IC:")
    c.drawString(50, 700, "Amount:")
    c.drawString(50, 670, "Product:")
    c.drawString(50, 640, "Date:")
    c.drawString(50, 610, "Staff:")
    c.drawString(50, 580, "RM Code:")
    c.drawString(50, 550, "Branch:")
    c.save()


def _write_json(path: Path, data) -> None:
    path.write_text(json.dumps(data, indent=2), encoding="utf-8")


def test_package_engine_generates_ordered_combined_pdf() -> None:
    with tempfile.TemporaryDirectory() as td:
        root = Path(td)
        forms_root = root / "forms"
        output_root = root / "output"
        fixture_root = root / "json"

        _make_pdf(forms_root / "first_source.pdf", "FIRST_SOURCE_TEMPLATE")
        _make_pdf(forms_root / "second_source.pdf", "SECOND_SOURCE_TEMPLATE")
        fixture_root.mkdir()

        _write_json(
            fixture_root / "procedures.json",
            [
                {
                    "ProcedureCode": "PROC_SMOKE",
                    "DisplayName": "Two Form Smoke",
                    "Active": True,
                    "AutoBlankAfterOdd": False,
                }
            ],
        )
        _write_json(
            fixture_root / "source_forms.json",
            [
                {
                    "SourceFormCode": "SF_FIRST",
                    "DisplayName": "First Source",
                    "PDFFilePath": "first_source.pdf",
                    "MappingKey": "first_mapping",
                    "Active": True,
                },
                {
                    "SourceFormCode": "SF_SECOND",
                    "DisplayName": "Second Source",
                    "PDFFilePath": "second_source.pdf",
                    "MappingKey": "second_mapping",
                    "Active": True,
                },
            ],
        )
        _write_json(
            fixture_root / "procedure_items.json",
            [
                {
                    "ProcedureCode": "PROC_SMOKE",
                    "StepNo": 3,
                    "ItemType": "SourceForm",
                    "SourceFormCode": "SF_SECOND",
                    "BlankPageCount": 0,
                },
                {
                    "ProcedureCode": "PROC_SMOKE",
                    "StepNo": 1,
                    "ItemType": "SourceForm",
                    "SourceFormCode": "SF_FIRST",
                    "BlankPageCount": 0,
                },
                {
                    "ProcedureCode": "PROC_SMOKE",
                    "StepNo": 2,
                    "ItemType": "BlankPage",
                    "SourceFormCode": "",
                    "BlankPageCount": 1,
                },
            ],
        )
        _write_json(
            fixture_root / "forms.json",
            {
                "_shared_fields": {},
                "first_mapping": {
                    "name": "First Mapping",
                    "fields": [
                        {
                            "name": "name",
                            "ExcelColumnOrField": "name",
                            "source": "data",
                            "page": 1,
                            "x": 100,
                            "y": 760,
                            "required": True,
                        },
                        {
                            "name": "amount",
                            "ExcelColumnOrField": "amount",
                            "source": "data",
                            "page": 1,
                            "x": 100,
                            "y": 700,
                            "required": True,
                            "format": "currency_2_decimals",
                        },
                        {
                            "name": "date",
                            "source": "session",
                            "session_key": "date",
                            "page": 1,
                            "x": 100,
                            "y": 640,
                        },
                        {
                            "name": "staff_name",
                            "source": "staff_profile",
                            "profile_key": "staff_name",
                            "page": 1,
                            "x": 100,
                            "y": 610,
                        },
                        {
                            "name": "rm_branch",
                            "source": "session",
                            "session_key": "rm_branch",
                            "page": 1,
                            "x": 100,
                            "y": 550,
                        },
                    ],
                },
                "second_mapping": {
                    "name": "Second Mapping",
                    "fields": [
                        {
                            "name": "ic_number",
                            "ExcelColumnOrField": "ic_number",
                            "source": "data",
                            "page": 1,
                            "x": 100,
                            "y": 730,
                            "required": True,
                        },
                        {
                            "name": "product_type",
                            "ExcelColumnOrField": "product_type",
                            "source": "data",
                            "page": 1,
                            "x": 100,
                            "y": 670,
                        },
                        {
                            "name": "date",
                            "source": "session",
                            "session_key": "date",
                            "page": 1,
                            "x": 100,
                            "y": 640,
                        },
                        {
                            "name": "staff_rm_code",
                            "source": "session",
                            "session_key": "staff_rm_code",
                            "page": 1,
                            "x": 100,
                            "y": 580,
                        },
                    ],
                },
            },
        )

        procedure = json.loads((fixture_root / "procedures.json").read_text(encoding="utf-8"))[0]
        source_forms = {
            row["SourceFormCode"]: row
            for row in json.loads((fixture_root / "source_forms.json").read_text(encoding="utf-8"))
        }
        procedure_items = json.loads((fixture_root / "procedure_items.json").read_text(encoding="utf-8"))
        forms_config = json.loads((fixture_root / "forms.json").read_text(encoding="utf-8"))

        customer_row = {
            "name": "Avery Tan",
            "ic_number": "900101101234",
            "cis": "CIS-987654",
            "branch_code": "KLG",
            "amount": "25000",
            "product_type": "UT Subscription",
        }

        result = package_engine.generate_package(
            procedure=procedure,
            source_forms=source_forms,
            procedure_items=procedure_items,
            forms_config=forms_config,
            client=customer_row,
            settings={
                "forms_folder": str(forms_root),
                "output_folder": str(output_root),
                "default_font_size": 10,
                "_staff_profile": {"staff_name": "CK Staff"},
            },
            output_root=output_root,
            session={"date": "19/05/2026", "staff_rm_code": "RM001"},
        )

        output_path = result["output_path"]
        assert result["status"] == "Success"
        assert result["warnings"] == []
        assert output_path.exists()
        assert list(output_root.glob("*.pdf")) == [output_path]
        assert "CIS" not in output_path.name.upper()
        assert customer_row["cis"] not in output_path.name

        reader = pypdf.PdfReader(str(output_path))
        assert len(reader.pages) == 3
        page_texts = [(page.extract_text() or "") for page in reader.pages]

        assert "FIRST_SOURCE_TEMPLATE" in page_texts[0]
        assert "Avery Tan" in page_texts[0]
        assert "25,000.00" in page_texts[0]
        assert "19/05/2026" in page_texts[0]
        assert "CK Staff" in page_texts[0]
        assert "KLG" not in page_texts[0]
        assert page_texts[1].strip() == ""
        assert "SECOND_SOURCE_TEMPLATE" in page_texts[2]
        assert "900101101234" in page_texts[2]
        assert "UT Subscription" in page_texts[2]
        assert "19/05/2026" in page_texts[2]
        assert "RM001" in page_texts[2]


def test_load_staff_profile_from_excel_table() -> None:
    with tempfile.TemporaryDirectory() as td:
        workbook_path = Path(td) / "clients.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Staff_Profile"
        ws.append([
            "staff_name",
            "staff_ic",
            "staff_id",
            "fimm_id",
            "ippc_id",
            "staff_position",
            "staff_rm_codes",
            "staff_branches",
        ])
        ws.append([
            "CK Staff",
            "900101101234",
            "S001",
            "FIMM123",
            "IPPC456",
            "Relationship Manager",
            "RM001, RM002",
            "KJG, HQ",
        ])
        wb.save(workbook_path)

        profile = excel_reader.load_staff_profile(workbook_path)

        assert profile["staff_name"] == "CK Staff"
        assert profile["rm_name"] == "CK Staff"
        assert profile["staff_ic"] == "900101101234"
        assert profile["staff_id"] == "S001"
        assert profile["fimm_id"] == "FIMM123"
        assert profile["ippc_id"] == "IPPC456"
        assert profile["staff_position"] == "Relationship Manager"
        assert profile["staff_rm_codes"] == ["RM001", "RM002"]
        assert profile["staff_branches"] == ["KJG", "HQ"]


def test_blank_staff_profile_falls_back_to_rm_profile() -> None:
    with tempfile.TemporaryDirectory() as td:
        workbook_path = Path(td) / "clients.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Staff_Profile"
        ws.append(["field", "value"])
        ws.append(["staff_name", ""])
        ws.append(["staff_id", ""])
        rm = wb.create_sheet("RM_Profile")
        rm.append(["field", "value"])
        rm.append(["rm_name", "Legacy RM"])
        rm.append(["staff_id", "L001"])
        rm.append(["branches", "KJG, PJY"])
        wb.save(workbook_path)

        profile = excel_reader.load_staff_profile(workbook_path)

        assert profile["staff_name"] == "Legacy RM"
        assert profile["staff_id"] == "L001"
        assert profile["staff_branches"] == ["KJG", "PJY"]


def test_workbook_schema_defaults_and_accounts() -> None:
    with tempfile.TemporaryDirectory() as td:
        workbook_path = Path(td) / "clients.xlsx"
        wb = Workbook()
        clients = wb.active
        clients.title = "default_clients"
        clients.append(["*cis", "*name", "*ic_number", "phone"])
        clients.append(["CIS001", "Avery Tan", "900101101234", "012"])
        accounts = wb.create_sheet("Default Accounts")
        accounts.append([
            "common_name", "account_type", "account_number",
            "holder_1_name", "holder_1_cis", "holder_1_ic",
            "holder_2_name", "holder_2_cis", "holder_2_ic",
        ])
        accounts.append([
            "Avery UT Wife", "UT", "UT123",
            "Avery Tan", "CIS001", "900101101234",
            "Mary Lee", "CIS002", "850202105678",
        ])
        staff = wb.create_sheet("default_staff")
        staff.append(["staff_name", "staff_rm_codes"])
        staff.append(["CK Staff", "RM001"])
        wb.save(workbook_path)

        schema = excel_reader.workbook_schema(workbook_path)
        default_clients = next(s for s in schema if s["sheet"] == "default_clients")
        defaults = {f["field"] for f in default_clients["fields"] if f["default"]}
        assert defaults == {"cis", "name", "ic_number"}

        records = excel_reader.load_customer_records(workbook_path)
        accounts = excel_reader.load_accounts(workbook_path)
        assert len(records) == 1
        rows = excel_reader.accounts_for_customer(accounts, records[0], "UT")
        assert len(rows) == 1
        assert rows[0]["common_name"] == "Avery UT Wife"
        assert "Avery UT Wife" in rows[0]["_label"]

        result = excel_reader.update_customer_field(
            workbook_path,
            records[0],
            "default_clients",
            "phone",
            "0199999999",
        )
        assert result["sheet"] == "default_clients"
        refreshed = excel_reader.load_customer_records(workbook_path)
        assert refreshed[0]["phone"] == "0199999999"


def test_sheet_aware_fields_and_manual_values() -> None:
    source_forms = {
        "SF_INV": {"SourceFormCode": "SF_INV", "MappingKey": "investment", "Active": True},
        "SF_INS": {"SourceFormCode": "SF_INS", "MappingKey": "insurance", "Active": True},
    }
    items = [
        {"ProcedureCode": "PROC", "StepNo": 1, "ItemType": "SourceForm", "SourceFormCode": "SF_INV"},
        {"ProcedureCode": "PROC", "StepNo": 2, "ItemType": "SourceForm", "SourceFormCode": "SF_INS"},
    ]
    forms = {
        "investment": {
            "fields": [{
                "name": "amount",
                "source": "data",
                "excel_sheet": "default_investment",
                "data_key": "amount",
                "required": True,
            }]
        },
        "insurance": {
            "fields": [{
                "name": "amount",
                "source": "data",
                "excel_sheet": "default_insurance",
                "data_key": "amount",
                "required": True,
            }]
        },
    }
    client = {
        "name": "Avery Tan",
        "_sheet_data": {
            "default_investment": {"amount": "1000"},
            "default_insurance": {"amount": ""},
        },
    }

    fields = package_engine.data_fields_for_procedure("PROC", forms, source_forms, items, include_common=True)
    assert {f["id"] for f in fields} == {"default_investment::amount", "default_insurance::amount"}
    missing = package_engine.missing_required_fields("PROC", client, forms, source_forms, items)
    assert [f["id"] for f in missing] == ["default_insurance::amount"]

    merged = package_engine.merge_manual_values(client, {"default_insurance::amount": "2000"})
    missing_after_manual = package_engine.missing_required_fields("PROC", merged, forms, source_forms, items)
    assert missing_after_manual == []

    both_missing = {
        "name": "Avery Tan",
        "_sheet_data": {
            "default_investment": {"amount": ""},
            "default_insurance": {"amount": ""},
        },
    }
    merged_one_sheet = package_engine.merge_manual_values(both_missing, {"default_insurance::amount": "2000"})
    still_missing = package_engine.missing_required_fields("PROC", merged_one_sheet, forms, source_forms, items)
    assert [f["id"] for f in still_missing] == ["default_investment::amount"]


def test_auto_blank_after_odd_pages() -> None:
    with tempfile.TemporaryDirectory() as td:
        root = Path(td)
        forms_root = root / "forms"
        output_root = root / "output"
        one = forms_root / "one.pdf"
        two = forms_root / "two.pdf"
        _make_pdf(one, "ONE_PAGE")
        _make_pdf(two, "TWO_PAGE")

        result = package_engine.generate_package(
            procedure={"ProcedureCode": "PROC_AUTO_BLANK", "DisplayName": "Auto Blank", "Active": True, "AutoBlankAfterOdd": True},
            source_forms={
                "SF_ONE": {"SourceFormCode": "SF_ONE", "PDFFilePath": "one.pdf", "MappingKey": "one", "Active": True},
                "SF_TWO": {"SourceFormCode": "SF_TWO", "PDFFilePath": "two.pdf", "MappingKey": "two", "Active": True},
            },
            procedure_items=[
                {"ProcedureCode": "PROC_AUTO_BLANK", "StepNo": 1, "ItemType": "SourceForm", "SourceFormCode": "SF_ONE"},
                {"ProcedureCode": "PROC_AUTO_BLANK", "StepNo": 2, "ItemType": "SourceForm", "SourceFormCode": "SF_TWO"},
            ],
            forms_config={"one": {"fields": []}, "two": {"fields": []}},
            client={"name": "Avery Tan"},
            settings={"forms_folder": str(forms_root), "default_font_size": 10},
            output_root=output_root,
            session={"date": "19/05/2026"},
        )

        reader = pypdf.PdfReader(str(result["output_path"]))
        assert len(reader.pages) == 4
        assert (reader.pages[1].extract_text() or "").strip() == ""
        assert (reader.pages[3].extract_text() or "").strip() == ""


def test_auto_blank_does_not_double_insert_before_manual_blank() -> None:
    with tempfile.TemporaryDirectory() as td:
        root = Path(td)
        forms_root = root / "forms"
        output_root = root / "output"
        _make_pdf(forms_root / "one.pdf", "ONE_PAGE")

        result = package_engine.generate_package(
            procedure={"ProcedureCode": "PROC_AUTO_MANUAL", "DisplayName": "Auto Manual", "Active": True, "AutoBlankAfterOdd": True},
            source_forms={
                "SF_ONE": {"SourceFormCode": "SF_ONE", "PDFFilePath": "one.pdf", "MappingKey": "one", "Active": True},
            },
            procedure_items=[
                {"ProcedureCode": "PROC_AUTO_MANUAL", "StepNo": 1, "ItemType": "SourceForm", "SourceFormCode": "SF_ONE"},
                {"ProcedureCode": "PROC_AUTO_MANUAL", "StepNo": 2, "ItemType": "BlankPage", "BlankPageCount": 1},
            ],
            forms_config={"one": {"fields": []}},
            client={"name": "Avery Tan"},
            settings={"forms_folder": str(forms_root), "default_font_size": 10},
            output_root=output_root,
            session={"date": "19/05/2026"},
        )

        reader = pypdf.PdfReader(str(result["output_path"]))
        assert len(reader.pages) == 2
        assert (reader.pages[1].extract_text() or "").strip() == ""


def test_source_folder_requires_exactly_one_top_level_pdf() -> None:
    with tempfile.TemporaryDirectory() as td:
        root = Path(td)
        forms_root = root / "forms"
        folder = forms_root / "SF001 Sample"
        folder.mkdir(parents=True)
        _make_pdf(folder / "one.pdf", "ONE")
        _make_pdf(folder / "two.pdf", "TWO")

        source = {
            "SourceFormCode": "SF001",
            "PDFFilePath": "SF001 Sample",
            "Active": True,
        }
        problem = catalog.source_pdf_path_problem(source, {"forms_folder": str(forms_root)})
        assert "exactly one top-level PDF" in problem

        try:
            package_engine.generate_package(
                procedure={"ProcedureCode": "PROC", "DisplayName": "Bad Folder", "Active": True},
                source_forms={"SF001": source},
                procedure_items=[{"ProcedureCode": "PROC", "StepNo": 1, "ItemType": "SourceForm", "SourceFormCode": "SF001"}],
                forms_config={"SF001": {"fields": []}},
                client={"name": "Avery Tan"},
                settings={"forms_folder": str(forms_root), "default_font_size": 10},
                output_root=root / "output",
                session={"date": "19/05/2026"},
            )
        except FileNotFoundError as e:
            assert "exactly one top-level PDF" in str(e)
        else:
            raise AssertionError("Expected folder with multiple PDFs to fail")


def test_source_form_date_issues() -> None:
    source_forms = {
        "SF_OLD": {
            "SourceFormCode": "SF_OLD",
            "DisplayName": "Old Form",
            "Active": True,
            "ExpiryDate": "2026-05-01",
        },
        "SF_SOON": {
            "SourceFormCode": "SF_SOON",
            "DisplayName": "Soon Form",
            "Active": True,
            "ExpiryDate": "2026-05-30",
        },
    }
    items = [
        {"ProcedureCode": "PROC", "StepNo": 1, "ItemType": "SourceForm", "SourceFormCode": "SF_OLD"},
        {"ProcedureCode": "PROC", "StepNo": 2, "ItemType": "SourceForm", "SourceFormCode": "SF_SOON"},
    ]

    issues = catalog.source_form_date_issues(
        "PROC",
        source_forms,
        items,
        as_of=catalog.parse_catalog_date("2026-05-27"),
    )

    assert [i["level"] for i in issues] == ["block", "warn"]
    assert "expired" in issues[0]["message"]
    assert "expires" in issues[1]["message"]


def test_recent_history_payload_round_trip_and_legacy_fallback() -> None:
    with tempfile.TemporaryDirectory() as td:
        path = Path(td) / "HistoryLog.xlsx"
        result = {
            "output_path": Path(td) / "out.pdf",
            "warnings": [],
            "status": "Success",
            "client": {
                "name": "Avery Tan",
                "client_name": "Avery Tan",
                "cis": "CIS001",
                "ic_number": "900101101234",
                "amount": "25000",
                "product_type": "UT",
                "action_purpose": "Subscribe",
            },
            "procedure_code": "P010",
            "procedure_name": "UT Investment",
            "session": {"date": "27/05/2026", "staff_rm_code": "RM001", "rm_branch": "KLG"},
            "account": {"common_name": "Avery UT", "account_number": "UT001", "_label": "Avery UT | UT | UT001"},
            "manual_values": {
                "default_investment::amount": "25000",
                "default_investment::product_type": "UT",
            },
        }
        excel_reader.append_history_rows(path, [package_engine.history_row(result, "CK Staff")])
        excel_reader.append_history_rows(path, [{
            "GeneratedDateTime": "2026-05-26 10:00:00",
            "ClientName": "Avery Tan",
            "CIS": "CIS001",
            "ProcedureCode": "P004",
            "ProcedureName": "FD Bundle",
            "OutputFilePath": "",
            "Amount": "10000",
            "ProductType": "FD",
            "ActionPurpose": "Renewal",
            "Status": "Success",
        }])

        rows = excel_reader.load_recent_history(path, {"cis": "CIS001", "name": "Avery Tan"}, limit=10)

        assert len(rows) == 2
        assert rows[0]["ProcedureCode"] == "P004"
        assert rows[0]["_payload"]["manual_values"]["amount"] == "10000"
        assert rows[1]["ProcedureCode"] == "P010"
        assert rows[1]["_payload"]["manual_values"]["default_investment::amount"] == "25000"
        assert rows[1]["_payload"]["account"]["account_number"] == "UT001"


def main() -> None:
    test_package_engine_generates_ordered_combined_pdf()
    test_load_staff_profile_from_excel_table()
    test_blank_staff_profile_falls_back_to_rm_profile()
    test_workbook_schema_defaults_and_accounts()
    test_sheet_aware_fields_and_manual_values()
    test_auto_blank_after_odd_pages()
    test_auto_blank_does_not_double_insert_before_manual_blank()
    test_source_folder_requires_exactly_one_top_level_pdf()
    test_source_form_date_issues()
    test_recent_history_payload_round_trip_and_legacy_fallback()
    print("OK: procedure package smoke")


if __name__ == "__main__":
    main()
