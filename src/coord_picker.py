"""
CoordPicker — visual tool to map PDF form field coordinates.

Open a blank form PDF, click where each field's text should appear,
fill in the field details, then save to forms.json.

Output: PDF-space coordinates (bottom-left origin).
  y=0 = bottom of page.  A4 top = y≈842.
  Use these values in forms.json field definitions.
"""
import argparse
import sys
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
import config_loader
import excel_reader

try:
    import fitz          # PyMuPDF
    from PIL import Image, ImageTk
    HAS_DEPS = True
except ImportError:
    HAS_DEPS = False

ZOOM = 1.5  # display scale factor (1.5 = 150% of actual PDF size)

SYSTEM_SOURCE_LABELS = ["Session Input", "Fixed Text", "Auto Date"]
FORMAT_OPTS   = [
    "text",
    "ic_dashed",
    "currency_2_decimals",
    "currency_4_decimals",
    "currency_no_decimals",
    "date_ddmmyyyy",
    "date_27_May_2026",
    "all_uppercase",
    "1st_letter_case",
]
ALIGN_OPTS    = ["left", "center", "right"]
AUTO_TYPES    = ["date", "year", "month"]
SESSION_FIELDS = ["date", "staff_rm_code", "rm_code", "rm_branch"]


class CoordPickerApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Coord Picker — Form Field Mapper")
        self.geometry("1050x700")
        self.resizable(True, True)
        self.minsize(800, 560)
        self._set_window_icon()

        self._doc        = None
        self._page_num   = 1
        self._page_count = 0
        self._page_w     = 0.0
        self._page_h     = 0.0
        self._photo      = None
        self._fields: list[dict] = []
        self._selected_idx: int | None = None
        self._drag_mode: str | None = None
        self._drag_start: tuple[float, float] | None = None
        self._last_suggested_format = "text"
        self._schema: list[dict] = []
        self._field_choices: list[str] = []

        # Form identity vars
        self._form_id       = tk.StringVar()
        self._form_name_var = tk.StringVar()
        self._subfolder     = tk.StringVar()

        # Field entry vars
        self._click_x    = tk.DoubleVar(value=0.0)
        self._click_y    = tk.DoubleVar(value=0.0)
        self._fname      = tk.StringVar()
        self._display    = tk.StringVar()
        self._source     = tk.StringVar(value="data")
        self._fmt        = tk.StringVar(value="text")
        self._auto_type  = tk.StringVar(value="date")
        self._fixed_val  = tk.StringVar()
        self._font_size  = tk.IntVar(value=10)
        self._align      = tk.StringVar(value="left")
        self._max_width  = tk.StringVar()
        self._max_chars  = tk.StringVar()
        self._overflow   = tk.StringVar(value="shrink_font")
        self._required   = tk.BooleanVar(value=False)

        self._build_ui()
        self._apply_startup_args()

        if not HAS_DEPS:
            messagebox.showerror(
                "Missing libraries",
                "PyMuPDF and Pillow are required.\n"
                "pip install PyMuPDF Pillow"
            )

    def _set_window_icon(self):
        bases = []
        if getattr(sys, "frozen", False):
            bases.append(Path(getattr(sys, "_MEIPASS", Path(sys.executable).parent)))
            bases.append(Path(sys.executable).parent)
        bases.append(Path(__file__).resolve().parents[1])
        for base in bases:
            icon = base / "assets" / "app.ico"
            if icon.exists():
                try:
                    self.iconbitmap(default=str(icon))
                    return
                except tk.TclError:
                    continue

    # ── UI construction ───────────────────────────────────────────────────────

    def _build_ui(self):
        # Left control panel
        left = tk.Frame(self, width=310, bg="#f0f0f0")
        left.pack(side=tk.LEFT, fill=tk.Y, padx=6, pady=6)
        left.pack_propagate(False)

        def section(parent, title):
            tk.Label(parent, text=title, bg="#f0f0f0",
                     font=("", 9, "bold"), anchor="w").pack(fill=tk.X, pady=(8, 2))

        def lbl_entry(parent, text, var):
            tk.Label(parent, text=text, bg="#f0f0f0", anchor="w").pack(fill=tk.X)
            tk.Entry(parent, textvariable=var).pack(fill=tk.X, pady=(0, 3))

        # ── Form identity ──
        section(left, "1. Form identity")
        lbl_entry(left, "MappingKey / SourceFormCode:", self._form_id)
        lbl_entry(left, "Display name:", self._form_name_var)

        tk.Label(left, text="Template subfolder:", bg="#f0f0f0",
                 anchor="w").pack(fill=tk.X)
        sub_row = tk.Frame(left, bg="#f0f0f0")
        sub_row.pack(fill=tk.X, pady=(0, 3))
        self._cmb_subfolder = ttk.Combobox(sub_row, textvariable=self._subfolder,
                                            state="normal", width=22)
        self._cmb_subfolder.pack(side=tk.LEFT)
        self._cmb_subfolder.bind("<<ComboboxSelected>>",
                                  self._on_subfolder_selected)
        tk.Button(sub_row, text="↻", width=3,
                  command=self._refresh_subfolders).pack(side=tk.LEFT, padx=2)
        self._refresh_subfolders()

        tk.Button(left, text="Open PDF…",
                  command=self._open_pdf).pack(fill=tk.X, pady=(2, 4))

        # Page nav
        nav = tk.Frame(left, bg="#f0f0f0")
        nav.pack(fill=tk.X, pady=(0, 6))
        tk.Button(nav, text="◄", width=3,
                  command=self._prev_page).pack(side=tk.LEFT)
        self.lbl_page = tk.Label(nav, text="No PDF", bg="#f0f0f0", width=14)
        self.lbl_page.pack(side=tk.LEFT, expand=True)
        tk.Button(nav, text="►", width=3,
                  command=self._next_page).pack(side=tk.LEFT)

        ttk.Separator(left, orient="horizontal").pack(fill=tk.X, pady=4)

        # ── Coordinates display ──
        section(left, "2. Click on PDF → coordinates appear here")
        cfrm = tk.Frame(left, bg="#f0f0f0")
        cfrm.pack(fill=tk.X)
        tk.Label(cfrm, text="x:", bg="#f0f0f0").pack(side=tk.LEFT)
        tk.Label(cfrm, textvariable=self._click_x,
                 width=8, relief=tk.SUNKEN).pack(side=tk.LEFT, padx=2)
        tk.Label(cfrm, text="y:", bg="#f0f0f0").pack(side=tk.LEFT, padx=(6, 0))
        tk.Label(cfrm, textvariable=self._click_y,
                 width=8, relief=tk.SUNKEN).pack(side=tk.LEFT, padx=2)

        ttk.Separator(left, orient="horizontal").pack(fill=tk.X, pady=6)

        # ── Field details ──
        section(left, "3. Field details")
        tk.Label(left, text="Field ID / Excel field:", bg="#f0f0f0", anchor="w").pack(fill=tk.X)
        self._cmb_field = ttk.Combobox(left, textvariable=self._fname, state="normal")
        self._cmb_field.pack(fill=tk.X, pady=(0, 3))
        self._cmb_field.bind("<<ComboboxSelected>>", self._on_field_name_change)
        self._updating_controls = False
        self._fname.trace_add("write", lambda *_: self._on_field_name_change())
        lbl_entry(left, "Display label:", self._display)

        tk.Label(left, text="Source:", bg="#f0f0f0", anchor="w").pack(fill=tk.X)
        self._cmb_source = ttk.Combobox(left, textvariable=self._source,
                                        values=[], state="readonly")
        self._cmb_source.pack(fill=tk.X, pady=(0, 3))
        self._cmb_source.bind("<<ComboboxSelected>>", self._on_source_change)
        tk.Button(left, text="Reload Excel Fields",
                  command=self._load_excel_schema).pack(fill=tk.X, pady=(0, 4))

        # Conditional panels
        self._pnl_fixed = tk.Frame(left, bg="#f0f0f0")
        tk.Label(self._pnl_fixed, text="Fixed value:", bg="#f0f0f0",
                 anchor="w").pack(fill=tk.X)
        tk.Entry(self._pnl_fixed,
                 textvariable=self._fixed_val).pack(fill=tk.X, pady=(0, 3))

        self._pnl_auto = tk.Frame(left, bg="#f0f0f0")
        tk.Label(self._pnl_auto, text="Auto type:", bg="#f0f0f0",
                 anchor="w").pack(fill=tk.X)
        ttk.Combobox(self._pnl_auto, textvariable=self._auto_type,
                     values=AUTO_TYPES, state="readonly").pack(fill=tk.X, pady=(0, 3))

        tk.Label(left, text="Format (optional):", bg="#f0f0f0",
                 anchor="w").pack(fill=tk.X)
        self._cmb_format = ttk.Combobox(left, textvariable=self._fmt,
                                        values=FORMAT_OPTS, state="readonly")
        self._cmb_format.pack(fill=tk.X, pady=(0, 4))
        self._cmb_format.bind("<<ComboboxSelected>>", lambda _=None: setattr(self, "_last_suggested_format", self._fmt.get()))

        bot_row = tk.Frame(left, bg="#f0f0f0")
        bot_row.pack(fill=tk.X, pady=(0, 4))
        tk.Label(bot_row, text="Font size:", bg="#f0f0f0").pack(side=tk.LEFT)
        tk.Spinbox(bot_row, from_=6, to=24, textvariable=self._font_size,
                   width=4).pack(side=tk.LEFT, padx=4)
        tk.Checkbutton(bot_row, text="Required", variable=self._required,
                       bg="#f0f0f0").pack(side=tk.LEFT)

        adv_row = tk.Frame(left, bg="#f0f0f0")
        adv_row.pack(fill=tk.X, pady=(0, 4))
        tk.Label(adv_row, text="Align:", bg="#f0f0f0").pack(side=tk.LEFT)
        ttk.Combobox(adv_row, textvariable=self._align, values=ALIGN_OPTS,
                     state="readonly", width=7).pack(side=tk.LEFT, padx=4)
        tk.Label(adv_row, text="Box width:", bg="#f0f0f0").pack(side=tk.LEFT)
        tk.Entry(adv_row, textvariable=self._max_width,
                 width=6).pack(side=tk.LEFT, padx=4)

        limit_row = tk.Frame(left, bg="#f0f0f0")
        limit_row.pack(fill=tk.X, pady=(0, 4))
        tk.Label(limit_row, text="Max chars:", bg="#f0f0f0").pack(side=tk.LEFT)
        tk.Entry(limit_row, textvariable=self._max_chars, width=6).pack(side=tk.LEFT, padx=4)
        tk.Label(limit_row, text="Overflow:", bg="#f0f0f0").pack(side=tk.LEFT)
        ttk.Combobox(limit_row, textvariable=self._overflow,
                     values=["shrink_font", "cut_text", "warn"],
                     state="readonly", width=10).pack(side=tk.LEFT, padx=4)

        btn_row = tk.Frame(left, bg="#f0f0f0")
        btn_row.pack(fill=tk.X, pady=4)
        tk.Button(btn_row, text="Add Field", bg="#28a745", fg="white",
                  command=self._add_field).pack(side=tk.LEFT, padx=(0, 4))
        tk.Button(btn_row, text="Update",
                  command=self._update_selected).pack(side=tk.LEFT, padx=(0, 4))
        tk.Button(btn_row, text="Delete",
                  command=self._delete_selected).pack(side=tk.LEFT, padx=(0, 4))
        self.lbl_count = tk.Label(btn_row, text="  0 fields",
                                  bg="#f0f0f0", fg="gray")
        self.lbl_count.pack(side=tk.LEFT, padx=6)

        clear_row = tk.Frame(left, bg="#f0f0f0")
        clear_row.pack(fill=tk.X, pady=(0, 4))
        tk.Button(clear_row, text="Clear Page",
                  command=self._clear_page_fields).pack(side=tk.LEFT, padx=(0, 4))
        tk.Button(clear_row, text="Clear All",
                  command=self._clear_all_fields).pack(side=tk.LEFT)

        ttk.Separator(left, orient="horizontal").pack(fill=tk.X, pady=4)

        tk.Label(left, text="Fields recorded:", bg="#f0f0f0",
                 anchor="w").pack(fill=tk.X)
        self.lst_fields = tk.Listbox(left, height=9, font=("Courier", 7), bg="#ffffff")
        self.lst_fields.pack(fill=tk.BOTH, expand=True)
        self.lst_fields.bind("<<ListboxSelect>>", self._on_list_select)

        ttk.Separator(left, orient="horizontal").pack(fill=tk.X, pady=4)
        tk.Button(left, text="Save to forms.json",
                  bg="#1F4E79", fg="white", height=2,
                  command=self._save).pack(fill=tk.X)

        # Right PDF canvas
        right = tk.Frame(self)
        right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True,
                   padx=(0, 6), pady=6)
        self.canvas = tk.Canvas(right, bg="#888888", cursor="crosshair",
                                highlightthickness=0)
        sb_v = ttk.Scrollbar(right, orient="vertical",
                             command=self.canvas.yview)
        sb_h = ttk.Scrollbar(right, orient="horizontal",
                             command=self.canvas.xview)
        self.canvas.configure(yscrollcommand=sb_v.set,
                              xscrollcommand=sb_h.set)
        sb_v.pack(side=tk.RIGHT, fill=tk.Y)
        sb_h.pack(side=tk.BOTTOM, fill=tk.X)
        self.canvas.pack(fill=tk.BOTH, expand=True)
        self.canvas.bind("<Button-1>", self._on_mouse_down)
        self.canvas.bind("<B1-Motion>", self._on_mouse_drag)
        self.canvas.bind("<ButtonRelease-1>", self._on_mouse_up)
        self.canvas.bind("<MouseWheel>",
                         lambda e: self.canvas.yview_scroll(
                             int(-1*(e.delta/120)), "units"))

        self.lbl_status = tk.Label(self, text="  Open a PDF to start.",
                                   anchor="w", fg="gray")
        self.lbl_status.pack(fill=tk.X, padx=8, pady=(0, 4))
        self._load_excel_schema()

    # ── Subfolder auto-discovery ──────────────────────────────────────────────

    def _refresh_subfolders(self):
        """Scan C:\\Forms (from settings) and populate subfolder dropdown."""
        try:
            settings = config_loader.load_settings()
            folders  = config_loader.scan_forms_folder(settings)
            self._cmb_subfolder["values"] = folders
            if folders and not self._subfolder.get():
                self._subfolder.set(folders[0])
        except Exception:
            pass  # settings.json may not exist yet

    def _load_excel_schema(self):
        """Refresh Source/Excel field dropdowns from the current clients.xlsx."""
        try:
            settings = config_loader.load_settings()
            workbook = config_loader.customer_workbook_path(settings)
            self._schema = excel_reader.workbook_schema(workbook)
        except Exception:
            self._schema = []
        choices = [s["sheet"] for s in self._schema] + SYSTEM_SOURCE_LABELS
        self._cmb_source["values"] = choices
        if choices and self._source.get() not in choices:
            self._source.set(choices[0])
        elif not choices:
            self._source.set("Fixed Text")
            self._cmb_source["values"] = SYSTEM_SOURCE_LABELS
        self._on_source_change()
        self._status(f"Excel fields loaded: {len(self._schema)} sheet(s).")

    def _schema_for_source(self, source_label: str) -> dict | None:
        for sheet in self._schema:
            if sheet["sheet"] == source_label:
                return sheet
        return None

    def _stored_source_for_label(self, source_label: str) -> tuple[str, dict]:
        if source_label == "Session Input":
            return "session", {"session_key": self._fname.get().strip()}
        if source_label == "Fixed Text":
            return "fixed", {}
        if source_label == "Auto Date":
            return "auto", {}
        sheet = self._schema_for_source(source_label)
        if sheet and sheet["sheet_key"] in {"default_staff", "staff_profile", "rm_profile"}:
            return "staff_profile", {"profile_key": self._fname.get().strip()}
        if sheet:
            return "data", {
                "excel_sheet": sheet["sheet"],
                "data_key": self._fname.get().strip(),
            }
        return "data", {"data_key": self._fname.get().strip()}

    def _on_subfolder_selected(self, _=None):
        """When RM picks a subfolder, auto-set form_id and try to load existing config."""
        sub = self._subfolder.get()
        if not sub:
            return
        if not self._form_id.get():
            self._form_id.set(sub.lower().replace(" ", "_").replace("-", "_"))
        # Load existing form fields if already mapped
        try:
            forms = config_loader.load_forms()
            for fid, fcfg in forms.items():
                if fid.startswith("_"):
                    continue
                if fcfg.get("template_subfolder") == sub:
                    self._form_id.set(fid)
                    self._form_name_var.set(fcfg.get("name", ""))
                    self._fields = list(fcfg.get("fields", []))
                    self._refresh_list()
                    self._status(
                        f"Loaded existing mapping for '{fid}' "
                        f"({len(self._fields)} fields). "
                        "Open PDF to re-map or add fields.")
                    break
        except Exception:
            pass

    def _load_existing_form_mapping(self, fid: str) -> bool:
        """Load an existing mapping by MappingKey/SourceFormCode before editing."""
        if not fid:
            return False
        try:
            forms = config_loader.load_forms()
            fcfg = forms.get(fid)
            if not isinstance(fcfg, dict):
                return False
            self._form_id.set(fid)
            if fcfg.get("name"):
                self._form_name_var.set(fcfg.get("name", ""))
            if fcfg.get("template_subfolder"):
                self._subfolder.set(fcfg.get("template_subfolder", ""))
            self._fields = list(fcfg.get("fields", []))
            self._selected_idx = None
            self._refresh_list()
            self._status(f"Loaded existing mapping for '{fid}' ({len(self._fields)} fields).")
            return True
        except Exception:
            return False

    # ── PDF ───────────────────────────────────────────────────────────────────

    def _open_pdf(self):
        if not HAS_DEPS:
            return
        path = filedialog.askopenfilename(
            title="Open blank form PDF",
            filetypes=[("PDF", "*.pdf"), ("All", "*.*")]
        )
        if not path:
            return
        self._load_pdf_path(Path(path))

    def _load_pdf_path(self, path: Path):
        if not HAS_DEPS:
            return
        self._doc = fitz.open(str(path))
        self._pdf_path = Path(path)
        self._page_count = len(self._doc)
        self._page_num = 1
        if not self._subfolder.get():
            self._subfolder.set(path.parent.name)
        if not self._form_id.get():
            self._form_id.set(path.parent.name.lower().replace(" ", "_"))
        self._render()

    def _apply_startup_args(self):
        parser = argparse.ArgumentParser(add_help=False)
        parser.add_argument("--form-id")
        parser.add_argument("--display-name")
        parser.add_argument("--subfolder")
        parser.add_argument("--pdf")
        args, _ = parser.parse_known_args()
        if args.form_id:
            self._form_id.set(args.form_id)
        if args.display_name:
            self._form_name_var.set(args.display_name)
        if args.subfolder:
            self._subfolder.set(args.subfolder)
        if args.form_id:
            self._load_existing_form_mapping(args.form_id)
            if args.display_name and not self._form_name_var.get().strip():
                self._form_name_var.set(args.display_name)
        if args.pdf:
            pdf_path = Path(args.pdf)
            if pdf_path.exists():
                self._load_pdf_path(pdf_path)

    def _render(self):
        if not self._doc:
            return
        page = self._doc[self._page_num - 1]
        self._page_w = page.rect.width
        self._page_h = page.rect.height
        pix = page.get_pixmap(matrix=fitz.Matrix(ZOOM, ZOOM))
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        self._photo = ImageTk.PhotoImage(img)
        self.canvas.delete("all")
        self.canvas.create_image(0, 0, anchor="nw", image=self._photo)
        self.canvas.configure(scrollregion=(0, 0, pix.width, pix.height))
        for idx, f in enumerate(self._fields):
            if f["page"] == self._page_num:
                self._draw_field(idx, f)
        self.lbl_page.config(
            text=f"Page {self._page_num} / {self._page_count}")
        self._status(f"PDF: {self._page_w:.0f}×{self._page_h:.0f} pt  |  "
                     "Click to pick coordinate")

    def _prev_page(self):
        if self._page_num > 1:
            self._page_num -= 1
            self._render()

    def _next_page(self):
        if self._doc and self._page_num < self._page_count:
            self._page_num += 1
            self._render()

    # ── Mouse/editing ────────────────────────────────────────────────────────

    def _on_mouse_down(self, e):
        if not self._doc:
            return
        cx = self.canvas.canvasx(e.x)
        cy = self.canvas.canvasy(e.y)
        idx, mode = self._field_hit(cx, cy)
        if idx is not None:
            self._select_field(idx)
            self._drag_mode = mode
            self._drag_start = (cx, cy)
            return
        px = round(cx / ZOOM, 1)
        py = round(self._page_h - cy / ZOOM, 1)
        self._click_x.set(px)
        self._click_y.set(py)
        self._selected_idx = None
        self._render()
        self._dot(cx, cy, "·")
        self._status(f"x={px}  y={py}  (page {self._page_num})  — "
                     "enter field name and click Add Field")

    def _on_mouse_drag(self, e):
        if self._selected_idx is None or self._drag_mode not in {"move", "resize"}:
            return
        cx = self.canvas.canvasx(e.x)
        cy = self.canvas.canvasy(e.y)
        field = self._fields[self._selected_idx]
        if self._drag_mode == "move":
            field["x"] = round(cx / ZOOM, 1)
            field["y"] = round(self._page_h - cy / ZOOM, 1)
            self._click_x.set(field["x"])
            self._click_y.set(field["y"])
        elif self._drag_mode == "resize":
            new_width = max(5.0, round((cx / ZOOM) - float(field.get("x", 0)), 1))
            field["max_width"] = new_width
            self._max_width.set(str(new_width))
        self._render()

    def _on_mouse_up(self, _):
        self._drag_mode = None
        self._drag_start = None
        self._refresh_list()

    def _field_hit(self, cx: float, cy: float) -> tuple[int | None, str | None]:
        items = self.canvas.find_overlapping(cx - 4, cy - 4, cx + 4, cy + 4)
        for item in reversed(items):
            tags = self.canvas.gettags(item)
            for tag in tags:
                if tag.startswith("handle:"):
                    return int(tag.split(":", 1)[1]), "resize"
                if tag.startswith("field:"):
                    return int(tag.split(":", 1)[1]), "move"
        return None, None

    def _draw_field(self, idx: int, field: dict):
        cx = float(field["x"]) * ZOOM
        cy = (self._page_h - float(field["y"])) * ZOOM
        label = field.get("DisplayLabel") or field.get("name")
        width = field.get("max_width") or field.get("MaxWidth")
        selected = idx == self._selected_idx
        color = "#0066ff" if selected else "red"
        tags = ("marker", f"field:{idx}")
        if width:
            try:
                w = float(width) * ZOOM
                h = max(14, float(field.get("font_size", 10)) * ZOOM * 1.4)
                self.canvas.create_rectangle(cx, cy - h + 4, cx + w, cy + 4,
                                             outline=color, width=2, tags=tags)
                self.canvas.create_rectangle(cx + w - 4, cy - h / 2, cx + w + 4, cy - h / 2 + 8,
                                             outline=color, fill="white", width=2,
                                             tags=("marker", f"handle:{idx}", f"field:{idx}"))
            except (TypeError, ValueError):
                pass
        self._dot(cx, cy, label, color=color, tags=tags)

    def _dot(self, cx, cy, label, color="red", tags="marker"):
        r = 5
        self.canvas.create_oval(cx-r, cy-r, cx+r, cy+r,
                                 outline=color, width=2, tags=tags)
        self.canvas.create_text(cx+8, cy, text=label, fill=color,
                                 font=("", 7), anchor="w", tags=tags)

    # ── Field management ──────────────────────────────────────────────────────

    def _on_source_change(self, _=None):
        if getattr(self, "_updating_controls", False):
            return
        self._pnl_fixed.pack_forget()
        self._pnl_auto.pack_forget()
        src = self._source.get()
        if src == "Fixed Text":
            self._pnl_fixed.pack(fill=tk.X, after=self._pnl_auto)
            self._field_choices = []
        elif src == "Auto Date":
            self._pnl_auto.pack(fill=tk.X, after=self._pnl_fixed)
            self._field_choices = AUTO_TYPES
        elif src == "Session Input":
            self._field_choices = SESSION_FIELDS
        else:
            sheet = self._schema_for_source(src)
            self._field_choices = [f["field"] for f in sheet["fields"]] if sheet else []
        self._cmb_field["values"] = self._field_choices
        self._on_field_name_change()

    def _suggest_format(self, field_name: str) -> str:
        key = excel_reader.normalize_header(field_name)
        if any(token in key for token in ("ic", "nric", "new_ic")):
            return "ic_dashed"
        if any(token in key for token in ("amount", "premium", "price", "fee", "charges", "value")):
            return "currency_2_decimals"
        if any(token in key for token in ("unit", "nav", "rate")):
            return "currency_4_decimals"
        if any(token in key for token in ("date", "dob", "effective", "expiry")):
            return "date_ddmmyyyy"
        return "text"

    def _on_field_name_change(self, _=None):
        if getattr(self, "_updating_controls", False):
            return
        name = self._fname.get().strip()
        if not name:
            return
        suggestion = self._suggest_format(name)
        if self._fmt.get() in ("", "text", self._last_suggested_format):
            self._fmt.set(suggestion)
            self._last_suggested_format = suggestion
        if not self._display.get().strip():
            self._display.set(name.replace("_", " ").title())

    def _field_from_controls(self) -> dict | None:
        name = self._fname.get().strip()
        if not name:
            messagebox.showwarning("Missing", "Enter a field name first.")
            return None
        stored_source, extra = self._stored_source_for_label(self._source.get())
        field: dict = {
            "name":      name,
            "page":      self._page_num,
            "x":         self._click_x.get(),
            "y":         self._click_y.get(),
            "source":    stored_source,
            "font_size": self._font_size.get(),
            "format":    self._fmt.get() or "text",
        }
        field.update(extra)
        if self._display.get().strip():
            field["DisplayLabel"] = self._display.get().strip()
        if self._align.get() != "left":
            field["alignment"] = self._align.get()
        if self._max_width.get().strip():
            try:
                field["max_width"] = float(self._max_width.get().strip())
            except ValueError:
                messagebox.showwarning("Bad box width", "Box width must be a number.")
                return None
        if self._max_chars.get().strip():
            try:
                field["max_chars"] = int(self._max_chars.get().strip())
            except ValueError:
                messagebox.showwarning("Bad max chars", "Max chars must be a whole number.")
                return None
        if self._overflow.get() != "shrink_font":
            field["overflow"] = self._overflow.get()
        if self._required.get():
            field["required"] = True
        if stored_source == "fixed" and self._fixed_val.get().strip():
            field["value"] = self._fixed_val.get().strip()
        if stored_source == "auto" and self._auto_type.get():
            field["auto_type"] = self._auto_type.get()
        return field

    def _add_field(self):
        field = self._field_from_controls()
        if not field:
            return
        self._fields.append(field)
        self._selected_idx = len(self._fields) - 1
        self._render()
        self._refresh_list()
        self._status(f"Added '{field['name']}'  (total: {len(self._fields)} fields)")

    def _update_selected(self):
        if self._selected_idx is None or self._selected_idx >= len(self._fields):
            messagebox.showwarning("No selection", "Select a field marker or list row first.")
            return
        field = self._field_from_controls()
        if not field:
            return
        self._fields[self._selected_idx] = field
        self._refresh_list()
        self._render()
        self._status(f"Updated '{field['name']}'")

    def _delete_selected(self):
        if self._selected_idx is None or self._selected_idx >= len(self._fields):
            return
        removed = self._fields.pop(self._selected_idx)
        self._selected_idx = None
        self._refresh_list()
        self._render()
        self._status(f"Removed '{removed['name']}'")

    def _clear_page_fields(self):
        if not messagebox.askyesno("Clear page", f"Remove all mappings on page {self._page_num}?"):
            return
        self._fields = [f for f in self._fields if f.get("page") != self._page_num]
        self._selected_idx = None
        self._refresh_list()
        self._render()

    def _clear_all_fields(self):
        if not messagebox.askyesno("Clear all", "Remove all mappings for this form?"):
            return
        self._fields = []
        self._selected_idx = None
        self._refresh_list()
        self._render()

    def _source_label_for_field(self, field: dict) -> str:
        src = field.get("source", "data")
        if src == "session":
            return "Session Input"
        if src == "fixed":
            return "Fixed Text"
        if src == "auto":
            return "Auto Date"
        if src in {"staff_profile", "rm_profile"}:
            for sheet in self._schema:
                if sheet["sheet_key"] in {"default_staff", "staff_profile", "rm_profile"}:
                    return sheet["sheet"]
            return "default_staff"
        return field.get("excel_sheet") or (self._schema[0]["sheet"] if self._schema else "data")

    def _load_field_to_controls(self, field: dict):
        self._updating_controls = True
        self._fname.set(field.get("data_key") or field.get("profile_key") or field.get("session_key") or field.get("name", ""))
        self._display.set(field.get("DisplayLabel", ""))
        source_label = self._source_label_for_field(field)
        if source_label not in self._cmb_source["values"]:
            values = list(self._cmb_source["values"])
            values.append(source_label)
            self._cmb_source["values"] = values
        self._source.set(source_label)
        self._fmt.set(field.get("format", "text") or "text")
        self._font_size.set(int(float(field.get("font_size", field.get("FontSize", 10)))))
        self._align.set(field.get("alignment", field.get("Alignment", "left")))
        self._max_width.set(str(field.get("max_width", field.get("MaxWidth", "")) or ""))
        self._max_chars.set(str(field.get("max_chars", field.get("MaxChars", "")) or ""))
        self._overflow.set(field.get("overflow", "shrink_font"))
        self._required.set(bool(field.get("required", field.get("Required", False))))
        self._fixed_val.set(field.get("value", field.get("DefaultValue", "")))
        self._auto_type.set(field.get("auto_type", "date"))
        self._click_x.set(float(field.get("x", field.get("X", 0))))
        self._click_y.set(float(field.get("y", field.get("Y", 0))))
        self._updating_controls = False
        self._on_source_change()

    def _select_field(self, idx: int):
        if idx < 0 or idx >= len(self._fields):
            return
        self._selected_idx = idx
        self._load_field_to_controls(self._fields[idx])
        self._render()
        self._refresh_list()
        self._status(f"Selected '{self._fields[idx].get('name', '')}'")

    def _on_list_select(self, _=None):
        sel = self.lst_fields.curselection()
        if sel:
            self._select_field(sel[0])

    def _refresh_list(self):
        self.lbl_count.config(text=f"  {len(self._fields)} fields")
        self.lst_fields.delete(0, tk.END)
        for f in self._fields:
            line = (f"p{f['page']}  "
                    f"x:{f['x']:<7.1f}  y:{f['y']:<7.1f}  "
                    f"{f['name']:<22} [{f['source']}]")
            if f.get("format"):
                line += f" {f['format']}"
            if f.get("required"):
                line += " *"
            self.lst_fields.insert(tk.END, line)
        if self._selected_idx is not None and self._selected_idx < len(self._fields):
            self.lst_fields.selection_clear(0, tk.END)
            self.lst_fields.selection_set(self._selected_idx)

    # ── Save ──────────────────────────────────────────────────────────────────

    def _save(self):
        fid = self._form_id.get().strip()
        if not fid:
            messagebox.showwarning("Missing", "Enter a Form ID.")
            return
        if not self._fields:
            messagebox.showwarning("Empty", "No fields added yet.")
            return
        try:
            existing = config_loader.load_forms()
        except Exception:
            existing = {}

        # Preserve internal blocks (_shared_fields, _TEMPLATE) — they hold
        # the shared field library + how-to-use docs. Only the form being
        # saved gets overwritten below.
        clean = dict(existing)
        pdf_filename = ""
        pdf_hash     = ""
        if self._doc and hasattr(self, "_pdf_path") and self._pdf_path:
            pdf_filename = self._pdf_path.name
            try:
                pdf_hash = config_loader.compute_template_hash(self._pdf_path)
            except Exception:
                pass

        clean[fid] = {
            "name":               self._form_name_var.get().strip() or fid,
            "template_subfolder": self._subfolder.get().strip(),
            "template_filename":  pdf_filename,
            "template_hash":      pdf_hash,
            "fields":             self._fields,
        }
        config_loader.save_forms(clean)
        template_path = None
        template_error = ""
        try:
            template_path = self._sync_excel_template(fid)
        except PermissionError:
            template_error = "\n\nExcel template was open/locked, so it was not updated."
        except Exception as e:
            template_error = f"\n\nExcel template was not updated: {e}"
        self._status(
            f"Saved {len(self._fields)} fields for '{fid}' → forms.json", "green")
        template_msg = (
            f"\n\nExcel template updated:\n{template_path}"
            if template_path else template_error
        )
        messagebox.showinfo(
            "Saved",
            f"'{fid}' saved to forms.json.\n"
            f"{len(self._fields)} field(s) across "
            f"{len({f['page'] for f in self._fields})} page(s)."
            f"{template_msg}"
        )

    def _sync_excel_template(self, form_id: str) -> Path | None:
        headers = []
        seen = set()
        for field in self._fields:
            if field.get("source", "data") != "data":
                continue
            name = field.get("name", "").strip()
            if not name or name in seen:
                continue
            seen.add(name)
            headers.append(name)

        if not headers:
            return None
        for required in reversed(["ic_number", "name"]):
            if required not in seen:
                headers.insert(0, required)
                seen.add(required)

        path = config_loader.data_path(f"{form_id}_template.xlsx")
        path.parent.mkdir(parents=True, exist_ok=True)

        if path.exists():
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            existing = [
                str(cell.value).strip()
                for cell in ws[1]
                if cell.value is not None and str(cell.value).strip()
            ]
            for header in headers:
                if header not in existing:
                    ws.cell(row=1, column=len(existing) + 1, value=header)
                    existing.append(header)
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Data"
            ws.append(headers)

        wb.save(path)
        wb.close()
        return path

    def _status(self, msg: str, color: str = "gray"):
        self.lbl_status.config(text=f"  {msg}", fg=color)


if __name__ == "__main__":
    CoordPickerApp().mainloop()
