"""
Excel reader — multi-sheet, fully dynamic.
Master sheet: client static info keyed by ic_number.
Batch sheets: transaction-specific data joined to Master by ic_number.

P1-1: All load_workbook() calls protected against PermissionError
       (Excel holds exclusive lock on Windows when file is open).
Fix-8: IC numbers normalized both sides of join.
Fix-9: Native Python types preserved through reader.
"""
import re
import shutil
import datetime
import json
import openpyxl
from pathlib import Path
from openpyxl.utils import get_column_letter


# ── IC normalization ──────────────────────────────────────────────────────────

def normalize_ic(raw) -> str:
    """Strip non-digits, zero-pad to 12. Empty → ''.
    Handles float from Excel (880101145678.0 → '880101145678').
    """
    # Convert float to int string first to avoid '880101145678.0'
    if isinstance(raw, float):
        raw = int(raw)
    cleaned = re.sub(r"[^0-9]", "", str(raw))
    return cleaned.zfill(12) if cleaned else ""


# ── Safe workbook open ────────────────────────────────────────────────────────

class ExcelLockedError(Exception):
    """Raised when Excel holds an exclusive lock on the file."""


def _open_wb(path: Path):
    """Open workbook, converting Windows PermissionError to ExcelLockedError."""
    try:
        return openpyxl.load_workbook(path, read_only=True, data_only=True)
    except PermissionError:
        raise ExcelLockedError(
            f"Cannot open '{path.name}' — it is locked by Excel.\n"
            "Close the file in Excel, then click ↻ Reload."
        )


def _open_wb_edit(path: Path):
    try:
        return openpyxl.load_workbook(
            path,
            keep_vba=Path(path).suffix.lower() == ".xlsm",
        )
    except PermissionError:
        raise ExcelLockedError(
            f"Cannot write '{path.name}' — it is locked by Excel.\n"
            "Close the file in Excel, then retry."
        )


# ── Sheet parsing ─────────────────────────────────────────────────────────────

def _sheet_to_dicts(ws) -> list[dict]:
    """
    Parse a worksheet. Native types (datetime, int, float) preserved.
    None cells → "". Strings stripped. ic_number auto-normalized.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [
        str(h).strip().lower().replace(" ", "_") if h else f"col_{i}"
        for i, h in enumerate(rows[0])
    ]
    result = []
    for row in rows[1:]:
        if not any(row):
            continue
        record = {}
        for k, v in zip(headers, row):
            if v is None:
                record[k] = ""
            elif isinstance(v, str):
                record[k] = v.strip()
            else:
                record[k] = v  # int, float, datetime preserved for formatters
        if "ic_number" in record:
            record["ic_number"] = normalize_ic(record["ic_number"])
        result.append(record)
    return result


# ── Generic customer workbook parsing ────────────────────────────────────────

_KEY_HEADERS = {
    "cis", "cis_no", "cis_number", "cif_no", "cif_number",
    "name", "client_name", "customer_name",
    "ic", "ic_number", "nric", "new_ic",
    "policy", "policy_no", "policy_number",
    "common_name", "account_number",
    "holder_1_cis", "holder_2_cis", "holder_3_cis",
    "holder_1_ic", "holder_2_ic", "holder_3_ic",
    "holder_1_name", "holder_2_name", "holder_3_name",
    "staff_name", "staff_ic", "staff_id", "fimm_id", "ippc_id",
    "staff_position", "staff_rm_codes", "staff_branches",
}

_CIS_ALIASES = ("cis", "cis_no", "cis_number", "cif_no", "cif_number")
_NAME_ALIASES = ("name", "client_name", "customer_name", "customer")
_IC_ALIASES = ("ic_number", "ic", "nric", "new_ic")
_POLICY_ALIASES = ("policy_number", "policy_no", "policy")


def normalize_header(raw) -> str:
    text = str(raw or "").strip().lower()
    if text.startswith("*"):
        text = text[1:].strip()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def is_default_header(raw) -> bool:
    return str(raw or "").strip().startswith("*")


def normalize_lookup_key(raw) -> str:
    return re.sub(r"[^a-zA-Z0-9]", "", str(raw or "")).upper()


def _detect_header_row(ws, max_scan: int = 20) -> int | None:
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), 1):
        headers = [normalize_header(v) for v in row if v not in (None, "")]
        if len(headers) < 2:
            continue
        if any(h in _KEY_HEADERS for h in headers):
            return idx
    return None


def _canonicalize_record(record: dict) -> dict:
    out = dict(record)

    for key in _CIS_ALIASES:
        if out.get(key) not in ("", None):
            out.setdefault("cis", str(out[key]).strip())
            if key == "cif_no":
                out.setdefault("cif_no", str(out[key]).strip())
            break

    for key in _NAME_ALIASES:
        if out.get(key) not in ("", None):
            out.setdefault("name", str(out[key]).strip())
            out.setdefault("client_name", str(out[key]).strip())
            break

    for key in _IC_ALIASES:
        if out.get(key) not in ("", None):
            out.setdefault("ic_number", normalize_ic(out[key]))
            break

    for key in _POLICY_ALIASES:
        if out.get(key) not in ("", None):
            out.setdefault("policy_number", str(out[key]).strip())
            break

    return out


def _sheet_records_generic(ws) -> list[dict]:
    header_row = _detect_header_row(ws)
    if not header_row:
        return []
    raw_headers = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    headers = [
        normalize_header(v) if v not in (None, "") else f"col_{i}"
        for i, v in enumerate(raw_headers)
    ]
    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(row):
            continue
        rec = {}
        for key, value in zip(headers, row):
            if value is None:
                rec[key] = ""
            elif isinstance(value, str):
                rec[key] = value.strip()
            else:
                rec[key] = value
        rows.append(_canonicalize_record(rec))
    return rows


def load_customer_records(path: Path) -> list[dict]:
    """
    Read a saved Excel workbook and merge customer rows across sheets.

    Matching priority is CIS, then IC, then policy number. Formula cells are read
    as their saved/calculated values through openpyxl data_only mode.
    """
    if not path.exists():
        return []
    wb = _open_wb(path)
    merged: dict[str, dict] = {}
    non_customer_sheets = {
        "rm_profile", "staff_profile", "default_staff",
        "leader_profile", "leaders",
        "bulk_cis_template", "history_log", "history",
        "default_accounts", "accounts",
    }
    for sheet_name in wb.sheetnames:
        if normalize_header(sheet_name) in non_customer_sheets:
            continue
        ws = wb[sheet_name]
        for rec in _sheet_records_generic(ws):
            key = ""
            if rec.get("cis"):
                key = "CIS:" + normalize_lookup_key(rec["cis"])
            elif rec.get("ic_number"):
                key = "IC:" + normalize_ic(rec["ic_number"])
            elif rec.get("policy_number"):
                key = "POL:" + normalize_lookup_key(rec["policy_number"])
            if not key:
                key = f"{sheet_name}:{len(merged) + 1}"
            existing = merged.setdefault(key, {})
            sheet_bucket = existing.setdefault("_sheet_data", {})
            sheet_bucket[sheet_name] = {k: v for k, v in rec.items() if v not in ("", None)}
            sheet_bucket[normalize_header(sheet_name)] = sheet_bucket[sheet_name]
            existing.update({k: v for k, v in rec.items() if v not in ("", None)})
            sheets = set(str(existing.get("_sheets", "")).split("|")) if existing.get("_sheets") else set()
            sheets.add(sheet_name)
            existing["_sheets"] = "|".join(sorted(sheets))
    wb.close()
    return sorted(merged.values(), key=lambda r: str(r.get("name") or r.get("cis") or ""))


def workbook_schema(path: Path) -> list[dict]:
    """
    Return sheet/field metadata for Mapping Editor dropdowns.

    A header starting with '*' is treated as a default/locked field marker.
    The returned field name removes the marker, so '*ic_number' becomes
    field='ic_number', default=True.
    """
    if not path.exists():
        return []
    wb = _open_wb(path)
    schema = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row = _detect_header_row(ws)
        if not header_row:
            continue
        raw_headers = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
        fields = []
        seen = set()
        for raw in raw_headers:
            if raw in (None, ""):
                continue
            name = normalize_header(raw)
            if not name or name in seen:
                continue
            seen.add(name)
            fields.append({
                "field": name,
                "display": str(raw).strip().lstrip("*").strip(),
                "default": is_default_header(raw),
            })
        if fields:
            schema.append({
                "sheet": sheet_name,
                "sheet_key": normalize_header(sheet_name),
                "fields": fields,
            })
    wb.close()
    return schema


def sheet_field_defaults(path: Path) -> dict[str, set[str]]:
    defaults: dict[str, set[str]] = {}
    for sheet in workbook_schema(path):
        fields = {f["field"] for f in sheet["fields"] if f.get("default")}
        if fields:
            defaults[sheet["sheet"]] = fields
            defaults[sheet["sheet_key"]] = fields
    return defaults


def _account_label(row: dict) -> str:
    common = row.get("common_name") or "Account"
    typ = row.get("account_type") or "-"
    num = row.get("account_number") or "-"
    holders = []
    for idx in range(1, 4):
        name = row.get(f"holder_{idx}_name")
        if name:
            holders.append(str(name))
    holder_text = " / ".join(holders)
    return f"{common} | {typ} | {num}" + (f" | {holder_text}" if holder_text else "")


def load_accounts(path: Path) -> list[dict]:
    """Load default_accounts rows. One row represents one selectable account."""
    if not path.exists():
        return []
    wb = _open_wb(path)
    sheet_name = None
    for candidate in ("default_accounts", "Default Accounts", "Accounts"):
        if candidate in wb.sheetnames:
            sheet_name = candidate
            break
    if not sheet_name:
        wb.close()
        return []
    rows = _sheet_records_generic(wb[sheet_name])
    wb.close()
    accounts = []
    for row in rows:
        if not (row.get("common_name") or row.get("account_number")):
            continue
        row["_source_sheet"] = sheet_name
        row["_label"] = _account_label(row)
        accounts.append(row)
    return accounts


def _account_search_values(account: dict) -> list[str]:
    values = [
        account.get("common_name", ""),
        account.get("account_type", ""),
        account.get("account_number", ""),
    ]
    for idx in range(1, 4):
        values.extend([
            account.get(f"holder_{idx}_name", ""),
            account.get(f"holder_{idx}_cis", ""),
            account.get(f"holder_{idx}_ic", ""),
        ])
    return [str(v) for v in values if v not in ("", None)]


def account_matches_customer(account: dict, customer: dict) -> bool:
    def compact_ic(raw) -> str:
        return re.sub(r"[^0-9]", "", str(raw or ""))

    customer_keys = {
        normalize_lookup_key(customer.get("cis", "")),
        normalize_lookup_key(customer.get("cif_no", "")),
        normalize_lookup_key(customer.get("name", "")),
        normalize_lookup_key(customer.get("client_name", "")),
    }
    customer_ics = {compact_ic(customer.get("ic_number", ""))}
    customer_keys.discard("")
    customer_ics.discard("")
    for value in _account_search_values(account):
        if normalize_lookup_key(value) in customer_keys or compact_ic(value) in customer_ics:
            return True
    return False


def accounts_for_customer(accounts: list[dict], customer: dict, account_type: str | None = None) -> list[dict]:
    rows = [a for a in accounts if account_matches_customer(a, customer)]
    if account_type:
        target = normalize_lookup_key(account_type)
        rows = [a for a in rows if normalize_lookup_key(a.get("account_type", "")) == target]
    return rows


def search_customers(records: list[dict], query: str, limit: int = 80) -> list[dict]:
    q = str(query or "").strip().lower()
    if not q:
        return records[:limit]
    q_key = normalize_lookup_key(q)
    out = []
    for rec in records:
        haystacks = [
            str(rec.get("name", "")).lower(),
            str(rec.get("client_name", "")).lower(),
            normalize_lookup_key(rec.get("cis", "")),
            normalize_lookup_key(rec.get("cif_no", "")),
            normalize_lookup_key(rec.get("ic_number", "")),
            normalize_lookup_key(rec.get("policy_number", "")),
        ]
        if any(q in h for h in haystacks[:2]) or any(q_key and q_key in h for h in haystacks[2:]):
            out.append(rec)
        if len(out) >= limit:
            break
    return out


def find_customer_by_cis(records: list[dict], cis: str) -> dict | None:
    target = normalize_lookup_key(cis)
    if not target:
        return None
    for rec in records:
        if normalize_lookup_key(rec.get("cis", "")) == target:
            return rec
        if normalize_lookup_key(rec.get("cif_no", "")) == target:
            return rec
    return None


def _cell_value_for_match(value, field: str) -> str:
    if field in {"ic", "ic_number", "nric", "new_ic", "holder_1_ic", "holder_2_ic", "holder_3_ic"}:
        return normalize_ic(value)
    return normalize_lookup_key(value)


def _sheet_context_for_customer(customer: dict, sheet_name: str) -> dict:
    sheet_data = customer.get("_sheet_data", {}) if isinstance(customer, dict) else {}
    context = dict(customer or {})
    if isinstance(sheet_data, dict):
        for key in (sheet_name, normalize_header(sheet_name), str(sheet_name).lower()):
            bucket = sheet_data.get(key)
            if isinstance(bucket, dict):
                context.update(bucket)
                break
    return context


def _row_matches_context(row: dict, context: dict) -> bool:
    strong_keys = [
        "account_number",
        "common_name",
        "cis",
        "cif_no",
        "cis_number",
        "ic_number",
        "ic",
        "nric",
        "new_ic",
        "policy_number",
        "policy_no",
        "policy",
    ]
    for key in strong_keys:
        if row.get(key) in ("", None) or context.get(key) in ("", None):
            continue
        if _cell_value_for_match(row.get(key), key) == _cell_value_for_match(context.get(key), key):
            return True
    for row_key, ctx_key in [
        ("holder_1_cis", "cis"),
        ("holder_2_cis", "cis"),
        ("holder_3_cis", "cis"),
        ("holder_1_ic", "ic_number"),
        ("holder_2_ic", "ic_number"),
        ("holder_3_ic", "ic_number"),
    ]:
        if row.get(row_key) in ("", None) or context.get(ctx_key) in ("", None):
            continue
        if _cell_value_for_match(row.get(row_key), row_key) == _cell_value_for_match(context.get(ctx_key), ctx_key):
            return True
    for key in ("name", "client_name", "customer_name"):
        if row.get(key) in ("", None) or context.get(key) in ("", None):
            continue
        if normalize_lookup_key(row.get(key)) == normalize_lookup_key(context.get(key)):
            return True
    return False


def update_customer_field(path: Path, customer: dict, sheet_name: str, field_name: str, value) -> dict:
    """
    Update one customer field in clients.xlsx and keep a timestamped backup.

    The target row is matched by account_number/common_name first when present,
    then CIS/IC/policy/name. This is intended for the UI Edit button beside
    default/locked fields.
    """
    if not path.exists():
        raise FileNotFoundError(path)
    wb = _open_wb_edit(path)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in {path.name}.")
        ws = wb[sheet_name]
        header_row = _detect_header_row(ws)
        if not header_row:
            raise ValueError(f"Cannot detect header row in sheet '{sheet_name}'.")

        raw_headers = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
        headers = [normalize_header(v) if v not in (None, "") else f"col_{i}" for i, v in enumerate(raw_headers)]
        field_key = normalize_header(field_name)
        if field_key not in headers:
            raise ValueError(f"Column '{field_name}' not found in sheet '{sheet_name}'.")
        target_col = headers.index(field_key) + 1

        context = _sheet_context_for_customer(customer, sheet_name)
        target_row = None
        for row_idx in range(header_row + 1, ws.max_row + 1):
            row = {}
            for col_idx, key in enumerate(headers, 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                row[key] = "" if cell_value is None else cell_value
            if _row_matches_context(_canonicalize_record(row), context):
                target_row = row_idx
                break
        if target_row is None:
            raise ValueError(f"Cannot find matching customer row in sheet '{sheet_name}'.")

        backup_dir = path.parent / "backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = backup_dir / f"{path.stem}_{stamp}{path.suffix}"
        if not backup_path.exists():
            shutil.copy(path, backup_path)

        ws.cell(row=target_row, column=target_col, value=value)
        wb.save(path)
        return {
            "sheet": sheet_name,
            "row": target_row,
            "column": target_col,
            "backup_path": str(backup_path),
        }
    finally:
        wb.close()


def read_cis_list(path: Path) -> list[str]:
    if not path.exists():
        return []
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        wb = _open_wb(path)
        ws = wb[wb.sheetnames[0]]
        vals = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell not in ("", None):
                    vals.append(str(cell).strip())
                    break
        wb.close()
        if vals and normalize_header(vals[0]) in {"cis", "cis_no", "cis_number"}:
            vals = vals[1:]
        return [v for v in vals if v]
    text = path.read_text(encoding="utf-8-sig")
    return [v.strip() for v in re.split(r"[\n,;]+", text) if v.strip()]


# ── Staff profile ───────────────────────────────────────────────────────────

_STAFF_PROFILE_KEYS = {
    "staff_name",
    "staff_ic",
    "staff_id",
    "fimm_id",
    "ippc_id",
    "staff_position",
    "position",
    "staff_rm_codes",
    "staff_branches",
    # Backward-compatible aliases from older RM_Profile sheet.
    "rm_name",
    "rm_ic",
    "rm_staff_id",
    "rm_code",
    "rm_codes",
    "branches",
}


def _split_csv(value) -> list[str]:
    return [v.strip() for v in str(value or "").split(",") if v.strip()]


def _profile_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _load_key_value_sheet(wb, sheet_name: str) -> dict:
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    out = {}
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] in (None, ""):
            continue
        key = normalize_header(row[0])
        if key in {"field", "key"}:
            continue
        out[key] = _profile_value(row[1] if len(row) >= 2 else "")
    return out


def _load_profile_sheet(wb, sheet_name: str) -> dict:
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    rows = [
        tuple(cell for cell in row)
        for row in ws.iter_rows(values_only=True)
        if any(cell not in (None, "") for cell in row)
    ]
    if not rows:
        return {}

    headers = [normalize_header(cell) for cell in rows[0]]
    recognized = [h for h in headers if h in _STAFF_PROFILE_KEYS]
    if len(recognized) >= 2:
        for row in rows[1:]:
            values = {}
            for idx, header in enumerate(headers):
                if header in _STAFF_PROFILE_KEYS:
                    values[header] = _profile_value(row[idx] if idx < len(row) else "")
            if any(values.values()):
                return values
        return {}

    values = {
        key: value
        for key, value in _load_key_value_sheet(wb, sheet_name).items()
        if key in _STAFF_PROFILE_KEYS
    }
    return values if any(values.values()) else {}


def load_staff_profile(path: Path) -> dict:
    """
    Load default staff information from default_staff / Staff_Profile.

    Backward-compatible: if default_staff / Staff_Profile is absent, RM_Profile
    is accepted. Returned aliases make mappings forgiving:
    staff_name/rm_name, staff_id/rm_staff_id, staff_rm_codes/rm_codes, etc.
    """
    profile = {k: "" for k in _STAFF_PROFILE_KEYS}
    profile["staff_rm_codes"] = []
    profile["staff_branches"] = []
    profile["rm_codes"] = []
    profile["branches"] = []
    if not path.exists():
        return profile

    wb = _open_wb(path)
    raw = _load_profile_sheet(wb, "default_staff")
    if not raw:
        raw = _load_profile_sheet(wb, "Staff_Profile")
    if not raw:
        raw = _load_profile_sheet(wb, "RM_Profile")
    wb.close()

    for key, value in raw.items():
        if key in profile:
            profile[key] = value

    # Normalize common aliases.
    profile["staff_name"] = profile.get("staff_name") or profile.get("rm_name", "")
    profile["rm_name"] = profile.get("rm_name") or profile.get("staff_name", "")
    profile["staff_ic"] = profile.get("staff_ic") or profile.get("rm_ic", "")
    profile["staff_id"] = profile.get("staff_id") or profile.get("rm_staff_id", "")
    profile["rm_staff_id"] = profile.get("rm_staff_id") or profile.get("staff_id", "")
    profile["staff_position"] = profile.get("staff_position") or profile.get("position", "")

    rm_codes = (
        profile.get("staff_rm_codes")
        or profile.get("rm_codes")
        or profile.get("rm_code")
        or ""
    )
    branches = profile.get("staff_branches") or profile.get("branches") or ""
    profile["staff_rm_codes"] = _split_csv(rm_codes)
    profile["rm_codes"] = profile["staff_rm_codes"]
    profile["staff_branches"] = _split_csv(branches)
    profile["branches"] = profile["staff_branches"]
    return profile


def load_rm_profile(path: Path) -> dict:
    """Backward-compatible alias used by older callers."""
    return load_staff_profile(path)


HISTORY_COLUMNS = [
    "GeneratedDateTime",
    "ClientName",
    "CIS",
    "ProcedureCode",
    "ProcedureName",
    "OutputFilePath",
    "Amount",
    "FDDetails",
    "ProductType",
    "ActionPurpose",
    "FollowUpNote",
    "Status",
    "ErrorMessage",
    "GeneratedBy",
]
HISTORY_PAYLOAD_COLUMN = "RestorePayload"
HISTORY_ALL_COLUMNS = HISTORY_COLUMNS + [HISTORY_PAYLOAD_COLUMN]


def _history_headers(ws) -> list[str]:
    row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
    return [str(c).strip() if c else "" for c in row]


def _ensure_history_headers(ws) -> list[str]:
    if ws.max_row < 1:
        ws.append(HISTORY_ALL_COLUMNS)
    existing = _history_headers(ws)
    if existing[:len(HISTORY_COLUMNS)] != HISTORY_COLUMNS:
        ws.insert_rows(1)
        for idx, col in enumerate(HISTORY_ALL_COLUMNS, 1):
            ws.cell(row=1, column=idx, value=col)
        existing = HISTORY_ALL_COLUMNS[:]
    if HISTORY_PAYLOAD_COLUMN not in existing:
        next_col = len(existing) + 1
        ws.cell(row=1, column=next_col, value=HISTORY_PAYLOAD_COLUMN)
        existing.append(HISTORY_PAYLOAD_COLUMN)
    payload_idx = existing.index(HISTORY_PAYLOAD_COLUMN) + 1
    ws.column_dimensions[get_column_letter(payload_idx)].hidden = True
    return existing


def append_history_rows(path: Path, rows: list[dict]) -> None:
    if not rows:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        wb = _open_wb_edit(path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "History"
        ws.append(HISTORY_ALL_COLUMNS)

    existing = _ensure_history_headers(ws)

    for row in rows:
        ws.append([row.get(col, "") for col in existing])
    wb.save(path)
    wb.close()


def _history_match_key(customer: dict) -> tuple[str, str, str]:
    cis = normalize_lookup_key(customer.get("cis") or customer.get("cif_no") or "")
    ic = normalize_lookup_key(customer.get("ic_number") or customer.get("ic") or "")
    name = normalize_lookup_key(customer.get("name") or customer.get("client_name") or "")
    return cis, ic, name


def _legacy_payload_from_history(row: dict) -> dict:
    manual_values = {}
    for col, key in [
        ("Amount", "amount"),
        ("FDDetails", "fd_details"),
        ("ProductType", "product_type"),
        ("ActionPurpose", "action_purpose"),
        ("FollowUpNote", "follow_up_note"),
    ]:
        if row.get(col) not in ("", None):
            manual_values[key] = row.get(col)
    return {
        "version": 0,
        "procedure_code": row.get("ProcedureCode", ""),
        "procedure_name": row.get("ProcedureName", ""),
        "manual_values": manual_values,
        "session": {},
        "account": {},
        "output_path": row.get("OutputFilePath", ""),
    }


def load_recent_history(path: Path, customer: dict, limit: int = 10) -> list[dict]:
    if not path.exists() or not customer:
        return []
    target_cis, target_ic, target_name = _history_match_key(customer)
    wb = _open_wb(path)
    ws = wb.active
    headers = _history_headers(ws)
    rows = []
    try:
        for values in reversed(list(ws.iter_rows(min_row=2, values_only=True))):
            row = {headers[idx]: (value if value is not None else "") for idx, value in enumerate(values) if idx < len(headers)}
            payload_raw = row.get(HISTORY_PAYLOAD_COLUMN, "")
            payload = {}
            if payload_raw:
                try:
                    payload = json.loads(str(payload_raw))
                except json.JSONDecodeError:
                    payload = {}

            payload_client = payload.get("client", {}) if isinstance(payload, dict) else {}
            row_cis = normalize_lookup_key(row.get("CIS") or payload_client.get("cis") or payload_client.get("cif_no") or "")
            row_ic = normalize_lookup_key(payload_client.get("ic_number") or payload_client.get("ic") or "")
            row_name = normalize_lookup_key(row.get("ClientName") or payload_client.get("name") or payload_client.get("client_name") or "")
            matches = (
                (target_cis and row_cis == target_cis)
                or (target_ic and row_ic == target_ic)
                or (not target_cis and not target_ic and target_name and row_name == target_name)
            )
            if not matches:
                continue
            if not payload:
                payload = _legacy_payload_from_history(row)
            row["_payload"] = payload
            rows.append(row)
            if len(rows) >= limit:
                break
    finally:
        wb.close()
    return rows


# ── Public loaders ────────────────────────────────────────────────────────────

def load_master(path: Path) -> dict[str, dict]:
    """Returns {normalized_ic: client_dict} from Master sheet."""
    wb = _open_wb(path)
    clients = _sheet_to_dicts(wb["Master"])
    wb.close()
    return {c["ic_number"]: c for c in clients if c.get("ic_number")}


def get_sheet_headers(path: Path, sheet_name: str) -> list[str]:
    """Column names from a batch sheet, excluding ic_number."""
    wb = _open_wb(path)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet_name]
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
    wb.close()
    return [
        str(h).strip().lower().replace(" ", "_")
        for h in first_row
        if h and str(h).strip().lower().replace(" ", "_") != "ic_number"
    ]


def load_batch(path: Path, sheet_name: str,
               master: dict[str, dict]) -> list[dict]:
    """Master + batch data merged by ic_number. Missing Master rows skipped."""
    wb = _open_wb(path)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(
            f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
        )
    rows = _sheet_to_dicts(wb[sheet_name])
    wb.close()
    merged = []
    for row in rows:
        ic = row.get("ic_number", "")
        master_rec = master.get(ic)
        if not master_rec:
            continue
        merged.append({**master_rec, **row})
    return merged


def find_client_in_batch(path: Path, sheet_name: str, ic_number: str) -> dict:
    """Return batch row for one client (normalized IC), or {} if not found."""
    ic_norm = normalize_ic(ic_number)
    try:
        wb = _open_wb(path)
    except ExcelLockedError:
        return {}
    if sheet_name not in wb.sheetnames:
        wb.close()
        return {}
    rows = _sheet_to_dicts(wb[sheet_name])
    wb.close()
    for row in rows:
        if row.get("ic_number") == ic_norm:
            return row
    return {}


def get_master_names(master: dict[str, dict]) -> list[str]:
    return sorted(c.get("name", "") for c in master.values() if c.get("name"))


def sheet_names(path: Path) -> list[str]:
    wb = openpyxl.load_workbook(path, read_only=True)
    names = wb.sheetnames
    wb.close()
    return names
